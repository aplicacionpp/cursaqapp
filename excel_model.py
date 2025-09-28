
from __future__ import annotations
from dataclasses import dataclass, asdict
from typing import Any, Dict, Iterable, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
import json
import pathlib

@dataclass
class CellInfo:
    sheet: str
    address: str
    value: Any
    formula: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        d = asdict(self)
        # Make JSON friendly
        if hasattr(self.value, "isoformat"):
            d["value"] = self.value.isoformat()
        return d

@dataclass
class SheetSummary:
    title: str
    max_row: int
    max_column: int
    nonempty_cells: int
    formula_count: int

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

class WorkbookModel:
    """
    Carga un libro de Excel y expone utilidades para explorar hojas, celdas y fórmulas.
    No recalcula fórmulas: muestra las fórmulas y, si existen, los valores en caché guardados por Excel.
    """
    def __init__(self, path: str) -> None:
        self.path = path
        self._wb_formula = None   # data_only=False -> permite leer fórmulas
        self._wb_values  = None   # data_only=True  -> permite leer valores cacheados
        self._load()

    def _load(self) -> None:
        self._wb_formula = load_workbook(self.path, data_only=False, read_only=False)
        self._wb_values  = load_workbook(self.path, data_only=True,  read_only=False)

    @property
    def sheet_names(self) -> List[str]:
        return list(self._wb_formula.sheetnames)

    def has_sheet(self, name: str) -> bool:
        return name in self._wb_formula.sheetnames

    def get_sheet(self, name: str) -> Worksheet:
        return self._wb_formula[name]

    def get_values_sheet(self, name: str) -> Worksheet:
        return self._wb_values[name]

    def sheet_summary(self, name: str) -> SheetSummary:
        ws_f = self.get_sheet(name)
        nonempty = 0
        formula_count = 0
        for row in ws_f.iter_rows():
            for c in row:
                if c.value not in (None, ""):
                    nonempty += 1
                    if isinstance(c.value, str) and c.value.startswith("="):
                        formula_count += 1
        return SheetSummary(
            title=name,
            max_row=ws_f.max_row,
            max_column=ws_f.max_column,
            nonempty_cells=nonempty,
            formula_count=formula_count,
        )

    def all_summaries(self) -> List[SheetSummary]:
        return [self.sheet_summary(n) for n in self.sheet_names]

    def iter_cells(self, name: str) -> Iterable[CellInfo]:
        ws_f = self.get_sheet(name)
        ws_v = self.get_values_sheet(name)
        for row in ws_f.iter_rows():
            for c in row:
                val = ws_v[c.coordinate].value
                formula = None
                if isinstance(c.value, str) and c.value.startswith("="):
                    formula = c.value
                if val is not None or formula is not None:
                    yield CellInfo(sheet=name, address=c.coordinate, value=val, formula=formula)

    def find_formulas(self, query: str = "") -> List[CellInfo]:
        """
        Busca celdas con fórmula; si query no está vacío, filtra por substring (case-insensitive).
        """
        results: List[CellInfo] = []
        q = (query or "").lower()
        for name in self.sheet_names:
            for info in self.iter_cells(name):
                if info.formula:
                    if not q or q in info.formula.lower():
                        results.append(info)
        return results

    def export_summary_json(self, out_path: str) -> str:
        data = {
            "workbook": str(self.path),
            "sheets": [s.to_dict() for s in self.all_summaries()],
        }
        pathlib.Path(out_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return out_path

    def dump_sheet_preview(self, name: str, max_rows: int = 20, max_cols: int = 20) -> List[List[Any]]:
        """
        Devuelve una vista previa de valores (no fórmulas) en forma de matriz.
        """
        ws_v = self.get_values_sheet(name)
        rmax = min(ws_v.max_row, max_rows)
        cmax = min(ws_v.max_column, max_cols)
        out: List[List[Any]] = []
        for r in ws_v.iter_rows(min_row=1, max_row=rmax, min_col=1, max_col=cmax, values_only=True):
            out.append(list(r))
        return out
